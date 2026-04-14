import glob
import os
import openpyxl
import pandas as pd

BASE = r'C:\Users\Usuário\vBox\ONE PAGE'

# Resolve os caminhos reais via glob para evitar problemas de encoding no terminal
def encontrar_arquivo(pasta, padrao):
    resultados = [f for f in glob.glob(os.path.join(pasta, padrao)) if not os.path.basename(f).startswith('~$')]
    return resultados[0] if resultados else None

ARQUIVO_1 = encontrar_arquivo(BASE, '*BASES ONE PAGE*.xlsx')
ARQUIVO_2 = encontrar_arquivo(BASE, 'ONE PAGE - ATUAL_V2.xlsm')

def inspecionar(caminho, engine='openpyxl'):
    print(f"\n{'='*60}")
    print(f"ARQUIVO: {caminho}")
    print(f"{'='*60}")
    if not caminho:
        print("ERRO: arquivo não encontrado!")
        return
    try:
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        print(f"Sheets disponíveis: {wb.sheetnames}")
        wb.close()

        for sheet in wb.sheetnames:
            if 'CAP' in sheet.upper():
                print(f"\n→ Inspecionando sheet: '{sheet}'")
                df = pd.read_excel(caminho, sheet_name=sheet, engine=engine, nrows=3)
                print(f"Colunas:")
                for col in df.columns:
                    print(f"  - '{col}'")
                print(f"\nPrimeiras linhas:")
                print(df.to_string())
    except Exception as e:
        print(f"ERRO: {e}")

inspecionar(ARQUIVO_1, engine='openpyxl')
inspecionar(ARQUIVO_2, engine='openpyxl')
